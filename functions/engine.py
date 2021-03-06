import os
from xlrd import *
from tkinter import *
from tkinter import ttk,  filedialog as filedia
from dotenv import load_dotenv
from pathlib import Path
from os import startfile
from tkinter import messagebox
from openpyxl import *
from calendar import monthlen
from datetime import datetime

#LOAD APP SETTINGS FROM THE .ENV FILE
env_path=Path('.')/'.env'

load_dotenv(dotenv_path=env_path)

def openDialog():
    myfile=filedia.askopenfilename(initialdir="./MyFiles")
    startfile(myfile)

def saveReceipt(receipt):
    file = os.open(os.getenv("FILENAME"),os.O_RDWR|os.O_CREAT)
    rec=str.encode(receipt.get("1.0","end"))
    os.write(file,rec)
    os.close(file)
    openDialog()

def clearScreen(receipt):
    receipt.delete("2.0", "end")

def daysCalculator(nodays,toPay,daysInMonth):
    if (nodays > 29):  # enable calculation in terms of months
        months = int(nodays / 31) + 1

        arrears = m_round(float(months * toPay),trac=0.05)

        return arrears
    else:  # enable calculation in terms of days where dates fall within the same month
        ndays = float(nodays / daysInMonth)

        arrears = m_round(float(ndays * toPay),trac=0.05)

        return arrears

def m_round(amnt,trac=0.05):#ENABLE PRECISION ON CENTS
      nmult = 0.05 * round(amnt / 0.05)

      return nmult

def  myValidator(dd):
        try:
            validNo=round((float(dd.get())),2)

            return validNo
        except ValueError:

            messagebox._show("ERROR","WRONG ENTRIES FOUND")


file = Path("./MyFiles/arrears.xlsx")

if file.exists():
    pass
else:
    file = Workbook()
    ws = file.active
    ws["A1"] = "DESCRIPTION"
    ws["B1"] = "NO OF DAYS"
    ws["C1"] = "W.E.F"
    ws["D1"] = "END DATE"
    ws["E1"] = "RIGHT PAY"
    ws["F1"] = "PAID AMOUNT"
    ws["G1"] = "ARREARS "
    file.save("./MyFiles/arrears.xlsx")


def genSheet(tarrears,nodays,startdate,enddate,rightPay,wrongPay,arrears):
     file=load_workbook("./MyFiles/arrears.xlsx")
     sheet=file.active
     dd=[str(tarrears),nodays,str(startdate),str(enddate),str(rightPay),str(wrongPay),str(arrears)]
     sheet.cell(column=1, row=sheet.max_row + 1,value=tarrears)
     sheet.cell(column=2,row=sheet.max_row,value=nodays)
     sheet.cell(column=3, row=sheet.max_row, value=startdate)
     sheet.cell(column=4, row=sheet.max_row, value=enddate)
     sheet.cell(column=5, row=sheet.max_row, value=rightPay)
     sheet.cell(column=6, row=sheet.max_row, value=wrongPay)
     sheet.cell(column=7, row=sheet.max_row, value=arrears)

     file.save("./MyFiles/arrears.xlsx")

def calculate(right_pay,paid_amount,arrears_choice,wef,end_date,receipt):
    # SANITIZE INPUTS
    getrightPay=myValidator(right_pay)
    getwrongPay=myValidator(paid_amount)
    tarrears=arrears_choice.get()

    rightPay=m_round(getrightPay,trac=0.05)
    wrongPay=m_round(getwrongPay,trac=0.05)

    toPay = float(rightPay - wrongPay)
    startdate=datetime.strptime(wef.get(),'%m/%d/%y').date()
    enddate = datetime.strptime(end_date.get(),'%m/%d/%y').date()

    nodays=abs(enddate-startdate).days +1

    daysInMonth = monthlen(enddate.year, enddate.month)

    arrears=daysCalculator(nodays,toPay,daysInMonth)

    receipt.insert(END, "\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\tARREARS\n")
    receipt.insert(END, "\t\t____________________________________________________________________________\n")
    receipt.insert(END, "\t\tDESCRIPTION:\t"+arrears_choice.get()+"\t\t\tDAYS TO PAY:\t"+str(nodays)+"\n\n")
    receipt.insert(END, "\t\tW.E.F:\t"+str(startdate)+"\t\t\tEND DATE:\t"+str(enddate)+"\n\n")
    receipt.insert(END, "\t\tRIGHT PAY:\t"+str(rightPay)+"\t\t\tAMOUNT PAID:\t"+str(wrongPay)+"\n")
    receipt.insert(END,"\t\t____________________________________________________________________________\n")
    receipt.insert(END, "\t\tARREARS\t"+ str(arrears)+"\n")
    receipt.insert(END,"\t\t--------------------------------------------------------------------------------------------\n")

    genSheet(tarrears, nodays, startdate, enddate, rightPay, wrongPay, arrears)

